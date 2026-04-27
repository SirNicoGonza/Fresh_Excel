import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DATE_FORMATS = [
    "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
    "%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d",
    "%m/%d/%Y", "%m-%d-%Y",
    "%d/%m/%y", "%d-%m-%y",
    "%Y%m%d",
    "%d %B %Y", "%d %b %Y",
    "%B %d, %Y", "%b %d, %Y",
]


def _try_parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(cleaned, fmt)
            except ValueError:
                continue
        try:
            return pd.to_datetime(cleaned, infer_datetime_format=True)
        except Exception:
            return None
    return None


def _fix_broken_chars(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    # Normalize unicode (fix mojibake-like issues)
    normalized = unicodedata.normalize("NFKC", value)
    # Remove non-printable control chars except \n \t
    cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", normalized)
    return cleaned.strip()


def _is_date_column(series: pd.Series) -> bool:
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    parsed = sum(1 for v in non_null if _try_parse_date(v) is not None)
    return parsed / len(non_null) >= 0.5


def _unmerge_workbook(wb_path: Path) -> tuple[Path, list[str]]:
    """Unmerge cells and fill with the original value. Returns path to fixed file."""
    logs = []
    wb = load_workbook(wb_path)

    for sheet in wb.worksheets:
        merged_ranges = list(sheet.merged_cells.ranges)
        if not merged_ranges:
            continue

        for merged_range in merged_ranges:
            top_left_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            sheet.unmerge_cells(str(merged_range))
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    sheet.cell(row, col).value = top_left_value
            logs.append(
                f"Hoja '{sheet.title}': celdas combinadas {merged_range} separadas y rellenadas con '{top_left_value}'"
            )

    out_path = wb_path.with_stem(wb_path.stem + "_unmerged")
    wb.save(out_path)
    return out_path, logs


def analyze_file(file_path: Path) -> dict:
    """Return a quick analysis of the file before cleaning."""
    wb = load_workbook(file_path)
    merged_count = sum(len(list(ws.merged_cells.ranges)) for ws in wb.worksheets)

    df_map = pd.read_excel(file_path, sheet_name=None, header=0)
    date_cols = []
    broken_char_cols = []
    total_rows = 0

    for sheet_name, df in df_map.items():
        total_rows += len(df)
        for col in df.columns:
            series = df[col].astype(str)
            if _is_date_column(df[col]):
                date_cols.append(f"{sheet_name}::{col}")
            if series.str.contains(r"[\x00-\x1f\x7f]|[\ufffd]", regex=True, na=False).any():
                broken_char_cols.append(f"{sheet_name}::{col}")

    return {
        "sheets": list(df_map.keys()),
        "total_rows": total_rows,
        "merged_cells": merged_count,
        "date_columns": date_cols,
        "broken_char_columns": broken_char_cols,
    }


def clean_file(
    file_path: Path,
    output_path: Path,
    fix_dates: bool = True,
    fix_broken_chars: bool = True,
    unmerge_cells: bool = True,
) -> dict:
    """
    Clean the Excel file and write result to output_path.
    Returns a report of all changes made.
    """
    report: dict[str, list[str]] = {"unmerge": [], "dates": [], "chars": [], "errors": []}
    working_path = file_path

    # --- Step 1: Unmerge cells via openpyxl ---
    if unmerge_cells:
        try:
            working_path, unmerge_logs = _unmerge_workbook(working_path)
            report["unmerge"] = unmerge_logs
        except Exception as e:
            report["errors"].append(f"Error al separar celdas combinadas: {e}")

    # --- Step 2: Load with pandas for data cleaning ---
    try:
        df_map: dict[str, pd.DataFrame] = pd.read_excel(working_path, sheet_name=None)
    except Exception as e:
        report["errors"].append(f"Error al leer el archivo: {e}")
        return report

    cleaned_map: dict[str, pd.DataFrame] = {}

    for sheet_name, df in df_map.items():
        # Fix broken characters
        if fix_broken_chars:
            for col in df.columns:
                if df[col].dtype == object:
                    original = df[col].copy()
                    df[col] = df[col].apply(_fix_broken_chars)
                    changed = (df[col] != original).sum()
                    if changed:
                        report["chars"].append(
                            f"Hoja '{sheet_name}', columna '{col}': {changed} celdas con caracteres rotos corregidas"
                        )

        # Fix date columns
        if fix_dates:
            for col in df.columns:
                if _is_date_column(df[col]):
                    original = df[col].copy()
                    df[col] = df[col].apply(lambda v: _try_parse_date(v) or v)
                    # Convert to consistent datetime dtype
                    try:
                        df[col] = pd.to_datetime(df[col], errors="coerce")
                    except Exception:
                        pass
                    changed = (df[col] != original).sum()
                    if changed:
                        report["dates"].append(
                            f"Hoja '{sheet_name}', columna '{col}': {changed} fechas normalizadas al formato YYYY-MM-DD"
                        )

        cleaned_map[sheet_name] = df

    # --- Step 3: Write output ---
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="YYYY-MM-DD") as writer:
            for sheet_name, df in cleaned_map.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Auto-fit column widths
        _autofit_columns(output_path)
    except Exception as e:
        report["errors"].append(f"Error al guardar el archivo limpio: {e}")

    # Cleanup temp unmerged file
    if working_path != file_path and working_path.exists():
        working_path.unlink()

    report["summary"] = {
        "unmerge_fixes": len(report["unmerge"]),
        "date_fixes": len(report["dates"]),
        "char_fixes": len(report["chars"]),
        "errors": len(report["errors"]),
    }

    return report


def _autofit_columns(file_path: Path) -> None:
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
    wb.save(file_path)
